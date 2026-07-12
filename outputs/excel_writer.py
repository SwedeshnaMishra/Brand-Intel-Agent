"""
Writes one row per brand into output.xlsx. Append-only: creates the file with
styled headers if it doesn't exist, otherwise adds a new row at the bottom.
"""

from __future__ import annotations
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet

from schemas import BrandIntelligence
from utils.logger import log

HEADERS = [
    "Brand Name",
    "Category",
    "Sub-Category",
    "Portals Live",
    "Check Ratings",       
    "Sellers Name",
    "Running Ads",
    "Listing Quality",
    "Weakness Report",
]


def _style_headers(ws: Worksheet) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = 24
    ws.freeze_panes = "A2"


def write_row(data: BrandIntelligence, path: str) -> None:
    try:
        if os.path.exists(path):
            wb = load_workbook(path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Brand Intelligence"
            _style_headers(ws)

        row = [
            data.brand_name,
            data.discovery.category or "N/A",
            data.discovery.sub_category or "N/A",
            ", ".join(data.portals_live) if data.discovery.found_on_amazon else "None found",
            "\n".join(data.marketplace.products_over_100_ratings) or "None",
            ", ".join(data.marketplace.sellers) or "None",
            "Yes" if data.marketplace.running_ads else "No",
            data.listing_quality.score,
            "\n".join(f"- {b}" for b in data.weakness.bullets),
        ]
        ws.append(row)

        new_row_idx = ws.max_row
        for col_idx in range(1, len(HEADERS) + 1):
            ws.cell(row=new_row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        wb.save(path)
        log("excel", f"saved row {new_row_idx} to {path}")
    except Exception as e:
        log("excel", f"failed to write row ({e})", ok=False)